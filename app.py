import streamlit as st
import pandas as pd
import numpy as np
import pydeck as pdk
import io
import base64
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for Degree Diploma Generator
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing, Circle, Rect, String

st.set_page_config(page_title="Silverfox Golf Tracker", layout="wide", page_icon="⛳")

# Custom UI Styling
st.markdown("""
    <style>
    .header-card {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        border-bottom: 3px solid #d97706;
        border-radius: 12px;
        padding: 24px 30px;
        margin-bottom: 25px;
        color: white;
    }
    .dark-silver-title {
        margin: 0;
        font-size: 32px;
        font-weight: 800;
        color: #94a3b8;
        letter-spacing: 0.5px;
    }
    .dark-silver-subtitle {
        margin: 6px 0 0 0;
        font-size: 18px;
        font-weight: 600;
        color: #64748b;
    }
    .metric-card {
        background-color: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 18px;
        text-align: center;
        box-shadow: 0 2px 4px rgba(0,0,0,0.02);
    }
    .metric-value {
        font-size: 28px;
        font-weight: 700;
        color: #0f172a;
    }
    .metric-label {
        font-size: 13px;
        font-weight: 600;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .diploma-card {
        background-color: #ffffff;
        border: 2px solid #d97706;
        border-radius: 12px;
        padding: 24px;
        color: #0f172a;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    }
    </style>
""", unsafe_allow_html=True)

# County Coordinates for Fallbacks
COUNTY_COORDS = {
    'Antrim': (54.85, -6.15), 'Armagh': (54.35, -6.65), 'Carlow': (52.70, -6.83),
    'Cavan': (53.98, -7.36), 'Clare': (52.87, -8.98), 'Cork': (51.89, -8.47),
    'Derry': (54.99, -6.90), 'Donegal': (54.80, -8.10), 'Down': (54.33, -5.92),
    'Dublin': (53.35, -6.26), 'Fermanagh': (54.35, -7.63), 'Galway': (53.27, -9.05),
    'Kerry': (52.14, -9.70), 'Kildare': (53.15, -6.81), 'Kilkenny': (52.65, -7.25),
    'Laois': (53.03, -7.30), 'Leitrim': (54.12, -8.00), 'Limerick': (52.66, -8.63),
    'Longford': (53.72, -7.80), 'Louth': (53.88, -6.40), 'Mayo': (53.80, -9.30),
    'Meath': (53.65, -6.68), 'Monaghan': (54.25, -6.97), 'Offaly': (53.23, -7.65),
    'Roscommon': (53.75, -8.20), 'Sligo': (54.27, -8.47), 'Tipperary': (52.60, -7.90),
    'Tyrone': (54.60, -7.10), 'Waterford': (52.15, -7.40), 'Westmeath': (53.53, -7.34),
    'Wexford': (52.33, -6.46), 'Wicklow': (53.00, -6.30)
}

# 1. Master Course List Loader using pre-computed CSV coordinates
@st.cache_data
def load_master_courses():
    try:
        df = pd.read_csv('Golf-course-list-Ireland-V2.csv')
    except Exception as e:
        st.error(f"Error loading course CSV file: {e}")
        return pd.DataFrame(columns=["Course ID", "Course Name", "County", "Province", "lat", "lon"])
    
    df.columns = [c.strip() for c in df.columns]
    rename_dict = {}
    for c in df.columns:
        c_lower = c.lower()
        if 'course' in c_lower or 'club' in c_lower:
            rename_dict[c] = 'Course Name'
        elif 'county' in c_lower:
            rename_dict[c] = 'County'
        elif 'province' in c_lower:
            rename_dict[c] = 'Province'
        elif c_lower in ['latitude', 'lat']:
            rename_dict[c] = 'lat'
        elif c_lower in ['longitude', 'lon', 'lng', 'long']:
            rename_dict[c] = 'lon'
            
    df = df.rename(columns=rename_dict)
    df['Course Name'] = df['Course Name'].astype(str).str.strip()
    df['County'] = df['County'].astype(str).str.strip()
    df['Province'] = df.get('Province', pd.Series(['Ireland'] * len(df))).astype(str).str.strip()
    df['Course ID'] = range(1, len(df) + 1)
    
    df['lat'] = pd.to_numeric(df.get('lat', None), errors='coerce')
    df['lon'] = pd.to_numeric(df.get('lon', None), errors='coerce')

    # County centroid fallback if any record is missing coordinates
    for idx, row in df.iterrows():
        if pd.isnull(row['lat']) or pd.isnull(row['lon']):
            fallback = COUNTY_COORDS.get(row['County'], (53.4, -7.8))
            df.at[idx, 'lat'] = fallback[0]
            df.at[idx, 'lon'] = fallback[1]

    return df

df_courses = load_master_courses()
total_courses = len(df_courses)

# 2. Pin Drop Icon Generator & SVG Logo
def create_pin_icon(fill_color):
    svg_pin = f'''<svg width="64" height="64" viewBox="0 0 64 64" xmlns="http://www.w3.org/2000/svg">
      <path d="M 32 2 C 20.95 2 12 10.95 12 22 C 12 35.5 32 60 32 60 C 32 60 52 35.5 52 22 C 52 10.95 43.05 2 32 2 Z" fill="{fill_color}" stroke="#ffffff" stroke-width="3"/>
      <circle cx="32" cy="22" r="8" fill="#ffffff"/>
    </svg>'''
    b64 = base64.b64encode(svg_pin.encode('utf-8')).decode('utf-8')
    return f"data:image/svg+xml;base64,{b64}"

GREEN_PIN_URL = create_pin_icon("#059669")
RED_PIN_URL = create_pin_icon("#dc2626")

def render_svg_as_image(svg_raw, width=120):
    clean_svg = svg_raw.strip().replace("\n", "").replace("\r", "")
    b64 = base64.b64encode(clean_svg.encode('utf-8')).decode('utf-8')
    st.image(f"data:image/svg+xml;base64,{b64}", width=width)

def get_silverfox_golf_logo():
    return '''<svg viewBox="0 0 180 180" xmlns="http://www.w3.org/2000/svg">
  <g>
    <!-- Background Crest Frame -->
    <circle cx="90" cy="90" r="82" fill="#0f172a" stroke="#d97706" stroke-width="4"/>
    <circle cx="90" cy="90" r="75" fill="none" stroke="#fef08a" stroke-width="1" stroke-dasharray="4,3"/>

    <!-- Bushy Silver Tail -->
    <path d="M 130,135 C 165,145 155,85 120,95 C 138,110 128,135 105,120 Z" fill="#64748b"/>
    <path d="M 142,108 C 158,102 152,90 135,94 C 144,98 140,105 130,105 Z" fill="#ffffff"/>

    <!-- Fox Body & Golf Shirt -->
    <path d="M 75,100 C 65,125 105,135 110,130 L 100,100 Z" fill="#1e293b"/>

    <!-- Silverfox Head -->
    <polygon points="90,45 60,75 100,72" fill="#64748b"/>
    <polygon points="60,75 42,78 65,86" fill="#ffffff"/>
    <polygon points="90,45 82,25 70,50" fill="#0f172a"/>
    <polygon points="87,45 82,29 74,48" fill="#f8fafc"/>
    <polygon points="100,52 100,30 86,55" fill="#0f172a"/>
    <polygon points="98,51 98,34 88,53" fill="#f8fafc"/>
    <ellipse cx="73" cy="66" rx="3" ry="4" fill="#0f172a"/>
    <circle cx="74" cy="65" r="1" fill="#ffffff"/>
    <polygon points="42,78 38,80 43,83" fill="#0f172a"/>

    <!-- Fox Arms Extended Holding Golf Club -->
    <path d="M 85,98 L 62,90" stroke="#64748b" stroke-width="6" stroke-linecap="round"/>
    <path d="M 92,98 L 66,92" stroke="#64748b" stroke-width="5" stroke-linecap="round"/>
    <ellipse cx="62" cy="90" rx="4" ry="4" fill="#cbd5e1"/>

    <!-- Golf Club (Shaft & Driver Head at Address) -->
    <line x1="62" y1="90" x2="48" y2="138" stroke="#e2e8f0" stroke-width="3.5" stroke-linecap="round"/>
    <path d="M 40,138 L 52,138 C 54,142 46,146 38,142 Z" fill="#f59e0b" stroke="#ffffff" stroke-width="1.5"/>

    <!-- Golf Ball on Tee -->
    <rect x="34" y="142" width="2" height="6" fill="#f59e0b"/>
    <circle cx="35" cy="139" r="3.5" fill="#ffffff" stroke="#94a3b8" stroke-width="0.5"/>
  </g>
</svg>'''

# 3. PDF Diploma Generator
def generate_pdf_diploma(played_count, total_count, completion_pct):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    styles = getSampleStyleSheet()
    
    header_eng = ParagraphStyle('CertEng', parent=styles['Normal'], fontName='Times-BoldItalic', fontSize=14, leading=18, textColor=colors.HexColor('#d97706'), alignment=1)
    title_style = ParagraphStyle('CertTitle', parent=styles['Normal'], fontName='Times-Bold', fontSize=24, leading=30, textColor=colors.HexColor('#0f172a'), alignment=1)
    subtitle_style = ParagraphStyle('CertSub', parent=styles['Normal'], fontName='Times-Italic', fontSize=13, leading=18, textColor=colors.HexColor('#475569'), alignment=1)
    body_style = ParagraphStyle('CertBody', parent=styles['Normal'], fontName='Times-Roman', fontSize=12, leading=18, textColor=colors.HexColor('#1e293b'), alignment=1)
    metrics_style = ParagraphStyle('CertMetrics', parent=styles['Normal'], fontName='Times-Bold', fontSize=16, leading=22, textColor=colors.HexColor('#059669'), alignment=1)

    seal_drawing = Drawing(80, 80)
    seal_drawing.add(Circle(40, 40, 36, fillColor=colors.HexColor('#d97706'), strokeColor=colors.HexColor('#b45309'), strokeWidth=2))
    seal_drawing.add(Circle(40, 40, 31, fillColor=None, strokeColor=colors.HexColor('#fef08a'), strokeWidth=1))
    seal_drawing.add(String(40, 42, "SILVERFOX", textAnchor='middle', fontName='Times-Bold', fontSize=8, fillColor=colors.HexColor('#ffffff')))
    seal_drawing.add(String(40, 32, "VERITAS", textAnchor='middle', fontName='Times-Bold', fontSize=7, fillColor=colors.HexColor('#fef08a')))

    story = [
        Spacer(1, 10),
        Paragraph("SILVERFOX ACADEMY OF GOLF EXCELLENCE", header_eng),
        Paragraph("<i>To All Who Shall See These Present Letters, Greetings</i>", subtitle_style),
        Spacer(1, 12),
        Paragraph("OFFICIAL DIPLOMA IN GOLFAHOLIC EXCELLENCE", title_style),
        Spacer(1, 12),
        Paragraph("By authority of the Governing Council, these presents certify that you are an official golfaholic, having demonstrated unwavering passion and dedication in pursuit of the Irish Links Quest.", body_style),
        Spacer(1, 14),
        Paragraph(f"Officially Certified Progress: <b>{played_count} OF {total_count} MASTER COURSES CONQUERED</b>", metrics_style),
        Paragraph(f"Official Completion Rate: <b>{completion_pct:.2f}%</b>", body_style),
        Spacer(1, 15),
        seal_drawing,
        Spacer(1, 15),
        Table([[Paragraph("__________________________<br/><b>Chancellor of Golf Ops</b>", body_style), Paragraph("__________________________<br/><b>Master of the Links</b>", body_style)]], colWidths=[300, 300])
    ]

    table = Table([[story]], colWidths=[720], rowHeights=[480])
    table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 4, colors.HexColor('#0f172a')),
        ('INNERGRID', (0,0), (-1,-1), 1.5, colors.HexColor('#d97706')),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#fffbeb')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 15),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
    ]))

    doc.build([table])
    buffer.seek(0)
    return buffer

# 4. OpenPyXL Excel Export Engine
def generate_excel_report(df_master, played_list):
    wb = openpyxl.Workbook()
    
    ws_stats = wb.active
    ws_stats.title = "Summary & Analytics"
    ws_stats.views.sheetView[0].showGridLines = True
    
    ws_list = wb.create_sheet(title="Master Course List")
    ws_list.views.sheetView[0].showGridLines = True
    
    fill_navy = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_slate = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_gold = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    fill_light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_played = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    
    f_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    f_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    f_label = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    f_value = Font(name="Segoe UI", size=11, color="0F172A")
    f_formula = Font(name="Segoe UI", size=11, bold=True, color="059669")

    thin_border = Border(
        left=Side(style='thin', color="CBD5E1"),
        right=Side(style='thin', color="CBD5E1"),
        top=Side(style='thin', color="CBD5E1"),
        bottom=Side(style='thin', color="CBD5E1")
    )

    ws_list.append(["Course ID", "Course Name", "County", "Province", "Status"])
    for col_num in range(1, 6):
        c = ws_list.cell(row=1, column=col_num)
        c.font = f_header
        c.fill = fill_navy
        c.alignment = Alignment(horizontal="center", vertical="center")

    for idx, row in df_master.iterrows():
        is_played = row["Course Name"] in played_list
        status_str = "PLAYED" if is_played else "UNPLAYED"
        ws_list.append([row["Course ID"], row["Course Name"], row["County"], row["Province"], status_str])
        
        row_idx = idx + 2
        for col_num in range(1, 6):
            cell = ws_list.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            cell.font = f_value
            if is_played:
                cell.fill = fill_played

    max_data_row = len(df_master) + 1

    ws_stats.merge_cells("A1:E2")
    ws_stats["A1"] = "SILVERFOX GOLF TRACKER - EXECUTIVE REPORT"
    ws_stats["A1"].font = f_title
    ws_stats["A1"].fill = fill_navy
    ws_stats["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_stats["D4"] = "Status"
    ws_stats["E4"] = "Course Count"
    ws_stats["D4"].font = f_header
    ws_stats["E4"].font = f_header
    ws_stats["D4"].fill = fill_slate
    ws_stats["E4"].fill = fill_slate
    
    ws_stats["D5"] = "Played"
    ws_stats["E5"] = f"=COUNTIF('Master Course List'!E2:E{max_data_row}, \"PLAYED\")"
    
    ws_stats["D6"] = "Remaining"
    ws_stats["E6"] = f"=COUNTIF('Master Course List'!E2:E{max_data_row}, \"UNPLAYED\")"
    
    for row_c in range(5, 7):
        ws_stats[f"D{row_c}"].border = thin_border
        ws_stats[f"E{row_c}"].border = thin_border
        ws_stats[f"D{row_c}"].font = f_label
        ws_stats[f"E{row_c}"].font = f_formula

    ws_stats["A4"] = "KPI Metric"
    ws_stats["B4"] = "Value (Dynamic Formula)"
    ws_stats["A4"].font = f_header
    ws_stats["B4"].font = f_header
    ws_stats["A4"].fill = fill_gold
    ws_stats["B4"].fill = fill_gold

    kpi_formulas = [
        ("Total Master Courses", f"=COUNTA('Master Course List'!A2:A{max_data_row})"),
        ("Courses Played", "=E5"),
        ("Remaining Courses", "=E6"),
        ("Completion Percentage", "=E5/B5")
    ]

    for r_idx, (label, formula) in enumerate(kpi_formulas, start=5):
        c_k = ws_stats[f"A{r_idx}"]
        c_v = ws_stats[f"B{r_idx}"]
        c_k.value = label
        c_v.value = formula
        c_k.font = f_label
        c_v.font = f_formula
        c_k.fill = fill_light
        c_v.fill = fill_light
        c_k.border = thin_border
        c_v.border = thin_border
        if r_idx == 8:
            c_v.number_format = '0.00%'

    pie = PieChart()
    pie.title = "Master Course Progress Breakdown"
    labels = Reference(ws_stats, min_col=4, min_row=5, max_row=6)
    data = Reference(ws_stats, min_col=5, min_row=4, max_row=6)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 18
    pie.height = 11

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True
    pie.dataLabels.showPercent = True
    pie.dataLabels.position = "outEnd"

    dp_played = DataPoint(idx=0)
    dp_played.graphicalProperties.solidFill = "059669"
    dp_remaining = DataPoint(idx=1)
    dp_remaining.graphicalProperties.solidFill = "64748B"
    pie.series[0].data_points = [dp_played, dp_remaining]
    
    ws_stats.add_chart(pie, "A11")

    for ws in [ws_stats, ws_list]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# 5. Header Section
st.markdown('<div class="header-card">', unsafe_allow_html=True)
col_h1, col_h2 = st.columns([1, 5])
with col_h1:
    render_svg_as_image(get_silverfox_golf_logo(), width=115)
with col_h2:
    st.markdown('<h1 class="dark-silver-title">Silverfox Golf Course Tracker</h1>', unsafe_allow_html=True)
    st.markdown(f'<p class="dark-silver-subtitle">Ireland\'s {total_courses} Master Golf Courses Quest ⛳</p>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# 6. Main Interactive Layout
col_map, col_controls = st.columns([1.5, 1])

with col_controls:
    st.markdown("### ⛳ Log Completed Courses")
    
    selected_courses = st.multiselect(
        "Search and select completed courses:",
        options=df_courses["Course Name"].tolist(),
        default=[c for c in ["Royal County Down Golf Club", "K Club (Palmer North)", "Portmarnock Golf Club"] if c in df_courses["Course Name"].values],
        key="golf_courses_selector"
    )

    num_played = len(selected_courses)
    completion_rate = (num_played / total_courses) * 100 if total_courses > 0 else 0

    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{total_courses}</div><div class="metric-label">Total</div></div>', unsafe_allow_html=True)
    with m2:
        st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#059669;">{num_played}</div><div class="metric-label">Played</div></div>', unsafe_allow_html=True)
    with m3:
        st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#d97706;">{completion_rate:.1f}%</div><div class="metric-label">Progress</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    
    st.markdown('<div class="diploma-card">', unsafe_allow_html=True)
    st.markdown("### 🎓 DIPLOMA IN GOLFAHOLIC EXCELLENCE")
    st.markdown(f"""
    **Silverfox Golfaholics Certification**
    
    Certify that you are an official golfaholic with **{num_played} OF {total_courses} COURSES** completed across Ireland (**{completion_rate:.2f}%** total completion).
    """)
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    excel_file = generate_excel_report(df_courses, selected_courses)
    pdf_diploma = generate_pdf_diploma(num_played, total_courses, completion_rate)
    
    b_col1, b_col2 = st.columns(2)
    with b_col1:
        st.download_button(
            label="🎓 Download Diploma (PDF)",
            data=pdf_diploma,
            file_name="silverfox_golfaholic_diploma.pdf",
            mime="application/pdf"
        )
    with b_col2:
        st.download_button(
            label="📊 Auditable Excel Report",
            data=excel_file,
            file_name="silverfox_golf_tracker_auditable.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with col_map:
    st.markdown("### 🗺️ Interactive Golf Map of Ireland")
    
    df_map = df_courses.copy()
    df_map["Played"] = df_map["Course Name"].isin(selected_courses)
    df_map["Status"] = df_map["Played"].map({True: "PLAYED", False: "UNPLAYED"})
    
    df_map["icon_data"] = df_map["Played"].apply(
        lambda p: {
            "url": GREEN_PIN_URL if p else RED_PIN_URL,
            "width": 64,
            "height": 64,
            "anchorY": 64
        }
    )

    icon_layer = pdk.Layer(
        "IconLayer",
        data=df_map,
        get_icon="icon_data",
        get_size=4,
        size_scale=8,
        get_position=["lon", "lat"],
        pickable=True
    )

    view_state = pdk.ViewState(
        latitude=53.4,
        longitude=-7.8,
        zoom=5.8,
        pitch=25
    )

    r = pdk.Deck(
        layers=[icon_layer],
        initial_view_state=view_state,
        tooltip={
            "html": "<b>{Course Name}</b><br/>County: {County}<br/>Province: {Province}<br/>Status: <b>{Status}</b>",
            "style": {"backgroundColor": "#0f172a", "color": "white", "borderRadius": "8px", "padding": "10px"}
        }
    )

    st.pydeck_chart(r, use_container_width=True)
    st.caption("🔴 **Red Pins:** Courses Remaining | 🟢 **Green Pins:** Courses Played")

st.divider()

# 7. Master Checklist Export
st.markdown("### 📋 Master Checklist Export")
df_export_master = df_courses.copy()
df_export_master["Played"] = df_export_master["Course Name"].isin(selected_courses)
df_export_master["Status"] = df_export_master["Played"].map({True: "PLAYED", False: "UNPLAYED"})

csv_master = df_export_master[["Course ID", "Course Name", "County", "Province", "Status"]].to_csv(index=False).encode('utf-8')

st.download_button(
    label="📥 Download Master Course List (CSV)",
    data=csv_master,
    file_name="silverfox_master_course_checklist.csv",
    mime="text/csv"
)